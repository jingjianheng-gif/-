#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
一号挤出机 e1_ 变量 Modbus 地址导出工具
=============================================
通过 TIA Portal Openness API 获取所有 e1_ 变量的 LogicalAddress，
换算为 Modbus 线圈/寄存器地址，输出 Excel + 自动更新 Markdown 清单。

前置条件：
  1. TIA Portal 已启动，项目已打开
  2. pythonnet 已安装：pip install pythonnet
  3. openpyxl 已安装：pip install openpyxl

用法:
    python export_e1_modbus.py

输出:
    导出结果/一号挤出机_e1_Modbus地址.xlsx    ← Excel 详细表
    设备说明/一号挤出机变量清单.md            ← 更新后的 Markdown（自动备份原文件）

作者: CodeWhale
日期: 2026-06-30
"""

import sys
import os
import re
import datetime
import traceback
from pathlib import Path

# ── Windows 控制台 UTF-8 ──
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# ── 路径配置 ──
SCRIPT_DIR = Path(__file__).parent.resolve()
OUTPUT_DIR = SCRIPT_DIR / "导出结果"
MARKDOWN_PATH = SCRIPT_DIR / "设备说明" / "一号挤出机变量清单.md"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# TIA Portal API 路径（按安装版本调整）
TIA_API_PATHS = {
    "V17": r"C:\Program Files\Siemens\Automation\Portal V17\PublicAPI\V17",
    "V19": r"C:\Program Files\Siemens\Automation\Portal V19\PublicAPI\V19",
}


# ═══════════════════════════════════════════════
#  Modbus 地址换算
# ═══════════════════════════════════════════════

def parse_logical_address(addr_str):
    """
    解析 TIA Portal LogicalAddress 字符串。
    返回: (area, byte_offset, bit_offset, size_hint)
      area: 'I','Q','M','MW','MD','IW','ID','DB','UDT'
      size_hint: 'bit','word','dword'
    示例:
      '%I0.0'  → ('I', 0, 0, 'bit')
      '%Q2.3'  → ('Q', 2, 3, 'bit')
      '%M10.7' → ('M', 10, 7, 'bit')
      '%MW20'  → ('MW', 20, 0, 'word')
      '%MD44'  → ('MD', 44, 0, 'dword')
      '%IW6'   → ('IW', 6, 0, 'word')
    """
    if not addr_str:
        return (None, None, None, None)

    s = str(addr_str).strip()

    # 位地址格式 %Ix.y, %Qx.y, %Mx.y
    m = re.match(r'^%([IQM])(\d+)\.(\d+)$', s)
    if m:
        area = m.group(1)
        byte_off = int(m.group(2))
        bit_off = int(m.group(3))
        return (area, byte_off, bit_off, 'bit')

    # 字/双字格式 %MWx, %MDx, %IWx, %IDx
    m = re.match(r'^%(MW|MD|IW|ID)(\d+)$', s)
    if m:
        area = m.group(1)
        offset = int(m.group(2))
        size = 'dword' if area in ('MD', 'ID') else 'word'
        return (area, offset, 0, size)

    # DB 块格式 %DBn.DBWx, %DBn.DBDx, %DBn.DBBx
    m = re.match(r'^%(DB\d+)\.DB([WDB])(\d+)$', s)
    if m:
        db = m.group(1)  # e.g. DB1
        typ = m.group(2)  # W, D, B
        offset = int(m.group(3))
        if typ == 'B':
            return (db, offset, 0, 'byte')
        elif typ == 'W':
            return (db, offset, 0, 'word')
        elif typ == 'D':
            return (db, offset, 0, 'dword')

    return (None, None, None, None)


def to_modbus_address(area, byte_offset, bit_offset, size_hint):
    """
    将解析后的逻辑地址转换为 Modbus 地址字符串。

    西门子 S7 标准 Modbus 映射：
      %Ix.y  → 离散输入  1xxxx  = 10001 + x*8 + y
      %Qx.y  → 线圈      0xxxx  =     1 + x*8 + y
      %Mx.y  → 线圈      0xxxx  =     1 + x*8 + y
      %MWx   → 保持寄存器 4xxxx  = 40001 + x
      %MDx   → 保持寄存器 4xxxx  = 40001 + x (2 regs)
      %IWx   → 输入寄存器 3xxxx  = 30001 + x
      %DBn.DBWx → 保持寄存器 4xxxx = 40001 + DB偏移

    返回: (modbus_5digit, modbus_area_name, fc_hint)
    """
    if area is None:
        return ("", "", "")

    if area == 'I':
        addr = 10001 + byte_offset * 8 + bit_offset
        return (str(addr), "离散输入 (1xxxx)", "FC02")
    elif area == 'Q':
        addr = 1 + byte_offset * 8 + bit_offset
        return (str(addr).zfill(5), "线圈 (0xxxx)", "FC01/05/15")
    elif area == 'M':
        addr = 1 + byte_offset * 8 + bit_offset
        return (str(addr).zfill(5), "线圈 (0xxxx)-M区", "FC01/05/15")
    elif area == 'MW':
        addr = 40001 + byte_offset
        return (str(addr), "保持寄存器 (4xxxx)", "FC03/06/16")
    elif area == 'MD':
        addr = 40001 + byte_offset
        return (str(addr), "保持寄存器 (4xxxx)-双字", "FC03/06/16 (2 regs)")
    elif area == 'IW':
        addr = 30001 + byte_offset
        return (str(addr), "输入寄存器 (3xxxx)", "FC04")
    elif area == 'ID':
        addr = 30001 + byte_offset
        return (str(addr), "输入寄存器 (3xxxx)-双字", "FC04 (2 regs)")
    elif area and area.startswith('DB'):
        if size_hint == 'byte':
            addr = 40001 + byte_offset
        elif size_hint == 'word':
            addr = 40001 + byte_offset // 2
        elif size_hint == 'dword':
            addr = 40001 + byte_offset // 2
        else:
            addr = 40001 + byte_offset
        return (str(addr), f"保持寄存器 (4xxxx)-{area}", "FC03/06/16")
    else:
        return ("", "", "")


# ═══════════════════════════════════════════════
#  TIA Portal 连接与变量导出
# ═══════════════════════════════════════════════

def find_openness_api():
    """查找已安装的 TIA Portal Openness API"""
    for version, path in TIA_API_PATHS.items():
        dll = Path(path) / "Siemens.Engineering.dll"
        if dll.exists():
            return version, path
    return None, None


def connect_and_export():
    """连接 TIA Portal 并导出 e1_ 变量"""
    version, api_path = find_openness_api()
    if version is None:
        print("[✗] 未找到 TIA Portal Openness API。")
        print("    请确认已安装 TIA Portal Professional/Advanced 并勾选 Openness 组件。")
        print(f"    已搜索: {list(TIA_API_PATHS.values())}")
        sys.exit(1)

    print(f"[✓] 找到 TIA Portal {version} Openness API: {api_path}")

    # ── 加载 Siemens.Engineering ──
    try:
        # pythonnet 2.x 方式
        import pythonnet
        pythonnet.load('netfx')
    except (ImportError, AttributeError):
        pass

    import clr
    import System

    dll_path = os.path.join(api_path, 'Siemens.Engineering.dll')
    System.Reflection.Assembly.LoadFrom(dll_path)
    clr.AddReference('Siemens.Engineering')

    from Siemens.Engineering import TiaPortal, TiaPortalMode

    print("\n连接 TIA Portal...")
    tia = TiaPortal(TiaPortalMode.WithUserInterface)
    projects = list(tia.Projects)
    if not projects:
        print("[✗] 未找到打开的项目。请在 TIA Portal 中打开项目后重试。")
        sys.exit(1)

    proj = projects[0]
    print(f"[✓] 项目: {proj.Path}")

    # ── 遍历所有设备，收集 e1_ 变量 ──
    all_tags = []
    for device in proj.Devices:
        dname = str(device.Name)
        for di in device.DeviceItems:
            try:
                sc = di.GetService('SoftwareContainer')
                if not sc or not sc.Software:
                    continue
                sw = sc.Software
                ttg = sw.TagTableGroup

                def walk(group, depth=0):
                    try:
                        for sg in group.Groups:
                            walk(sg, depth + 1)
                    except Exception:
                        pass
                    try:
                        for tt in group.TagTables:
                            tname = str(tt.Name)
                            for tag in tt.Tags:
                                tagname = str(tag.Name)
                                if not tagname.startswith('e1_'):
                                    continue
                                addr = str(tag.LogicalAddress) if hasattr(tag, 'LogicalAddress') else ''
                                dtype = str(tag.DataType) if hasattr(tag, 'DataType') else ''
                                comment = str(tag.Comment) if hasattr(tag, 'Comment') else ''
                                all_tags.append({
                                    'Name': tagname,
                                    'DataType': dtype,
                                    'LogicalAddress': addr,
                                    'Comment': comment,
                                    'TagTable': tname,
                                    'Device': dname,
                                })
                    except Exception as ex:
                        print(f"  [!] 读取变量表出错: {ex}")

                walk(ttg)
            except Exception:
                pass

    print(f"\n找到 {len(all_tags)} 个 e1_ 变量")

    if not all_tags:
        print("[✗] 未找到 e1_ 变量。请确认项目已打开且包含 PLC 变量表。")
        sys.exit(1)

    # ── 计算 Modbus 地址 ──
    for tag in all_tags:
        area, bo, bit, size = parse_logical_address(tag['LogicalAddress'])
        modbus_addr, modbus_area, fc = to_modbus_address(area, bo, bit, size)
        tag['ModbusAddress'] = modbus_addr
        tag['ModbusArea'] = modbus_area
        tag['FunctionCode'] = fc
        tag['ParsedArea'] = area if area else ''
        tag['ParsedOffset'] = str(bo) if bo is not None else ''

    return all_tags


# ═══════════════════════════════════════════════
#  输出 Excel
# ═══════════════════════════════════════════════

def write_excel(tags):
    """导出 Excel 文件"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "e1_Modbus地址"

    headers = [
        '变量名', '数据类型', '逻辑地址',
        'Modbus地址', 'Modbus区域', '功能码',
        '注释', '变量表', '设备'
    ]
    col_widths = [32, 14, 16, 16, 28, 18, 30, 20, 20]

    hf = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    h_align = Alignment(horizontal='center')

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hf
        cell.alignment = h_align

    # 按区域着色
    area_fills = {
        'I':  PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid'),  # 浅蓝
        'Q':  PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),  # 浅绿
        'M':  PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid'),  # 浅黄
        'MW': PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),  # 浅红
        'MD': PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),
        'IW': PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid'),  # 浅紫
        'ID': PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid'),
    }
    default_fill = PatternFill()

    for ri, tag in enumerate(tags, 2):
        row_data = [
            tag['Name'],
            tag['DataType'],
            tag['LogicalAddress'],
            tag['ModbusAddress'],
            tag['ModbusArea'],
            tag['FunctionCode'],
            tag['Comment'],
            tag['TagTable'],
            tag['Device'],
        ]
        area = tag.get('ParsedArea', '')
        fill = area_fills.get(area, default_fill)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill != default_fill:
                cell.fill = fill

    # 列宽
    for ci, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:I{len(tags) + 1}'

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"一号挤出机_e1_Modbus地址_{timestamp}.xlsx"
    wb.save(str(out_path))
    print(f"\n[✓] Excel 已保存: {out_path}")
    return out_path


# ═══════════════════════════════════════════════
#  更新 Markdown 清单
# ═══════════════════════════════════════════════

def classify_tag(tag):
    """
    根据变量名推断其所属类别。
    返回: (section_title, section_order)
    """
    name = tag['Name'].lower()

    # DO - 数字量输出
    if re.search(r'_do$|_do_spare$', name):
        return ('DO 数字量输出', 2)

    # DI - 数字量输入
    if re.search(r'_di$', name):
        return ('DI 数字量输入', 1)

    # AI - 模拟量输入
    if re.search(r'_ai$', name):
        return ('AI 模拟量输入', 3)

    # AQ - 模拟量输出
    if re.search(r'_aq$', name):
        return ('模拟量输出', 4)

    # DB 块参数
    db_patterns = [
        (r'^e1_connector_db\.', 'e1_connector_db'),
        (r'^e1_flange_db\.', 'e1_flange_db'),
        (r'^e1_watertemp_db\.', 'e1_watertemp_db'),
        (r'^e1_air_pump_db\.', 'e1_air_pump_db'),
        (r'^e1_alarm_db\.', 'e1_alarm_db'),
        (r'^e1_melt_hh_alarm', 'e1_melt_hh_alarm'),
    ]
    for pattern, db_name in db_patterns:
        if re.search(pattern, name):
            return (db_name, 50)

    # DB 块前缀
    if name.startswith('e1_') and '.' in name:
        db = name.split('.')[0]
        return (db, 50)

    # HMI 操作变量
    if 'hmi' in name:
        return ('HMI 操作变量', 60)

    # 其他
    return ('其他变量', 99)


def generate_markdown_table(tags_group, section_title):
    """
    为给定分组生成 Markdown 表格，含 Modbus 地址列。
    """
    lines = []
    lines.append(f"## {section_title}")
    lines.append("")
    lines.append("| 变量名 | 功能 | 数据类型 | 逻辑地址 | Modbus 地址 | Modbus 区域 |")
    lines.append("|--------|------|---------|---------|------------|------------|")
    for tag in tags_group:
        name = tag['Name']
        dtype = tag['DataType']
        laddr = tag['LogicalAddress']
        maddr = tag['ModbusAddress']
        marea = tag['ModbusArea']
        comment = tag['Comment'] or ''
        # 功能列：优先用注释，否则留空
        func = comment if comment else ''
        lines.append(f"| `{name}` | {func} | {dtype} | `{laddr}` | {maddr} | {marea} |")
    lines.append("")
    return "\n".join(lines)


def update_markdown(tags):
    """
    自动更新设备说明/一号挤出机变量清单.md。

    策略：
      1. 备份原文件为 .bak
      2. 重新生成整个 Markdown，保留原有的标题结构和 Modbus 说明
      3. 用 API 获取的数据替换所有变量表格（加入 Modbus 列）
    """
    if not MARKDOWN_PATH.exists():
        print(f"[!] 目标文件不存在: {MARKDOWN_PATH}")
        # 直接生成新文件
        return generate_new_markdown(tags)

    # 备份
    bak = MARKDOWN_PATH.with_suffix('.md.bak')
    import shutil
    shutil.copy2(str(MARKDOWN_PATH), str(bak))
    print(f"[✓] 已备份: {bak}")

    # 读取原文件，提取头部（变量表之前的内容）
    with open(MARKDOWN_PATH, 'r', encoding='utf-8') as f:
        original = f.read()

    # 找到第一个 ## 之前的所有内容作为头部
    header_end = original.find('\n## ')
    if header_end == -1:
        header_end = original.find('\n---\n')
    if header_end == -1:
        header = original.strip()
    else:
        header = original[:header_end].strip()

    # 分类标签
    classified = {}
    order_map = {}
    for tag in tags:
        cat, order = classify_tag(tag)
        order_map[cat] = order
        if cat not in classified:
            classified[cat] = []
        classified[cat].append(tag)

    # 按 order 排序分类
    sorted_cats = sorted(classified.keys(), key=lambda c: order_map.get(c, 99))

    # 生成新内容
    new_lines = []
    if header:
        new_lines.append(header)
    new_lines.append("")
    new_lines.append(f"> 最后更新: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    new_lines.append("> 数据来源: TIA Portal Openness API 自动导出")
    new_lines.append(f"> e1_ 变量总数: {len(tags)}")
    new_lines.append("")

    # Modbus 地址说明
    new_lines.append("---")
    new_lines.append("")
    new_lines.append("## Modbus 地址换算说明")
    new_lines.append("")
    new_lines.append("| 逻辑地址类型 | Modbus 区域 | 地址公式 | 功能码 |")
    new_lines.append("|------------|-----------|---------|-------|")
    new_lines.append("| `%Ix.y` | 离散输入 (1xxxx) | `10001 + x*8 + y` | FC02 |")
    new_lines.append("| `%Qx.y` | 线圈 (0xxxx) | `1 + x*8 + y` | FC01/05/15 |")
    new_lines.append("| `%Mx.y` | 线圈 (0xxxx)-M区 | `1 + x*8 + y` | FC01/05/15 |")
    new_lines.append("| `%MWx` | 保持寄存器 (4xxxx) | `40001 + x` | FC03/06/16 |")
    new_lines.append("| `%MDx` | 保持寄存器 (4xxxx)-双字 | `40001 + x` | FC03/06/16 (2 regs) |")
    new_lines.append("| `%IWx` | 输入寄存器 (3xxxx) | `30001 + x` | FC04 |")
    new_lines.append("| `%DBn.DBWx` | 保持寄存器 (4xxxx) | DB 基址 + 偏移 | FC03/06/16 |")
    new_lines.append("")
    new_lines.append("> ⚠️ 实际 Modbus 地址取决于网关设备（CoTrust CTH2-277PN / ADFweb HD67607）的地址映射配置。")
    new_lines.append("> 以上地址为西门子标准映射规则推算值，请以设备组态为准。")
    new_lines.append("")

    # 各分类表格
    for cat in sorted_cats:
        tags_in_cat = classified[cat]
        # 排序：按名称排序
        tags_in_cat.sort(key=lambda t: t['Name'])
        new_lines.append("---")
        new_lines.append("")
        # 确定序号前缀
        if cat in ('DI 数字量输入', 'DO 数字量输出', 'AI 模拟量输入', '模拟量输出'):
            prefix_map = {
                'DI 数字量输入': '一、',
                'DO 数字量输出': '二、',
                'AI 模拟量输入': '三、',
                '模拟量输出': '四、',
            }
            prefix = prefix_map.get(cat, '')
            new_lines.append(generate_markdown_table(tags_in_cat, f"{prefix}{cat} ({len(tags_in_cat)}个)"))
        elif cat in ('HMI 操作变量',):
            new_lines.append(generate_markdown_table(tags_in_cat, f"六、{cat} ({len(tags_in_cat)}个)"))
        else:
            new_lines.append(generate_markdown_table(tags_in_cat, f"{cat} ({len(tags_in_cat)}个)"))

    # 写入
    new_content = "\n".join(new_lines)
    with open(MARKDOWN_PATH, 'w', encoding='utf-8') as f:
        f.write(new_content)

    print(f"[✓] Markdown 已更新: {MARKDOWN_PATH}")
    return MARKDOWN_PATH


def generate_new_markdown(tags):
    """全新生成 Markdown 清单（当原文件不存在时）"""
    # 创建占位文件，让 update_markdown 能正常走备份+生成流程
    MARKDOWN_PATH.parent.mkdir(parents=True, exist_ok=True)
    MARKDOWN_PATH.write_text("# placeholder\n", encoding='utf-8')
    return update_markdown(tags)


# ═══════════════════════════════════════════════
#  统计汇总
# ═══════════════════════════════════════════════

def print_summary(tags):
    """打印变量汇总统计"""
    area_counts = {}
    for tag in tags:
        area = tag.get('ParsedArea', 'unknown')
        area_counts[area] = area_counts.get(area, 0) + 1

    print("\n" + "=" * 50)
    print("  导出汇总")
    print("=" * 50)
    print(f"  e1_ 变量总数: {len(tags)}")
    print(f"\n  按逻辑地址类型分布:")
    area_labels = {
        'I':  '%I (数字量输入)',
        'Q':  '%Q (数字量输出)',
        'M':  '%M (中间位)',
        'MW': '%MW (字)',
        'MD': '%MD (双字)',
        'IW': '%IW (输入字)',
        'ID': '%ID (输入双字)',
    }
    for area in ['I', 'Q', 'M', 'MW', 'MD', 'IW', 'ID']:
        if area in area_counts:
            label = area_labels.get(area, area)
            print(f"    {label}: {area_counts[area]} 个")
    # DB 块
    db_areas = {k: v for k, v in area_counts.items() if k and (k.startswith('DB') or k not in area_labels)}
    if db_areas:
        print(f"    DB 块: {sum(db_areas.values())} 个")
        for k, v in sorted(db_areas.items()):
            print(f"      {k}: {v} 个")

    # 无地址
    unknown = area_counts.get('unknown', 0)
    if unknown:
        print(f"    ⚠ 无逻辑地址: {unknown} 个")


# ═══════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  一号挤出机 e1_ 变量 Modbus 地址导出工具")
    print("  基于 TIA Portal Openness API")
    print("=" * 60)
    print()

    # 1. 连接 TIA Portal 并导出变量
    tags = connect_and_export()

    # 2. 统计
    print_summary(tags)

    # 3. 输出 Excel
    excel_path = write_excel(tags)

    # 4. 更新 Markdown
    md_path = update_markdown(tags)

    print()
    print("=" * 60)
    print("  ✅ 全部完成！")
    print(f"  Excel:    {excel_path}")
    print(f"  Markdown: {md_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
