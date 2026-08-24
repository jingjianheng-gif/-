#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
从剪贴板读取 TIA Portal 变量表，输出带 Modbus 地址的结果。

用法:
  1. 在 TIA Portal 中打开 PLC 变量表（如"默认变量表"）
  2. 筛选框输入 e1_ 回车
  3. Ctrl+A 全选 → Ctrl+C 复制
  4. 运行: python capture_clipboard.py
"""
import sys, os, re, datetime
from pathlib import Path
import subprocess

# Windows 控制台 UTF-8
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

SCRIPT_DIR = Path(__file__).parent.resolve()
OUTPUT_DIR = SCRIPT_DIR / "导出结果"
MARKDOWN_PATH = SCRIPT_DIR / "设备说明" / "一号挤出机变量清单.md"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── 1. 读剪贴板 ──
try:
    import win32clipboard
    win32clipboard.OpenClipboard()
    data = win32clipboard.GetClipboardData()
    win32clipboard.CloseClipboard()
except ImportError:
    try:
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        data = root.clipboard_get()
        root.destroy()
    except:
        # 最后尝试用 PowerShell
        import subprocess
        result = subprocess.run(['powershell', '-Command', 'Get-Clipboard'], 
                              capture_output=True, text=True, encoding='utf-8')
        data = result.stdout

if not data or len(data) < 10:
    print("[✗] 剪贴板为空或内容太少。请先在博图中 Ctrl+A → Ctrl+C 复制变量表。")
    sys.exit(1)

print(f"剪贴板内容: {len(data)} 字符")
print(f"前200字符: {data[:200]}")

# ── 2. 解析剪贴板数据 ──
# TIA Portal 导出的格式通常是 Tab 分隔的列
lines = [l.strip() for l in data.strip().split('\n') if l.strip()]
print(f"\n共 {len(lines)} 行")

# 自动检测列分隔符和结构
# 尝试 Tab 分隔
sample = lines[0]
if '\t' in sample:
    sep = '\t'
    headers = sample.split(sep)
    data_lines = lines[1:]
else:
    # 可能只是名称列表
    print("未检测到 Tab 分隔，假设是纯变量名列表")
    sep = None
    headers = ['Name']
    data_lines = lines

print(f"列: {headers}")
print(f"数据行: {len(data_lines)}")


# ── Modbus 地址换算函数 ──
def parse_logical_address(addr_str):
    if not addr_str:
        return (None, None, None, None)
    s = str(addr_str).strip()
    m = re.match(r'^%([IQM])(\d+)\.(\d+)$', s)
    if m:
        return (m.group(1), int(m.group(2)), int(m.group(3)), 'bit')
    m = re.match(r'^%(MW|MD|IW|ID)(\d+)$', s)
    if m:
        return (m.group(1), int(m.group(2)), 0, 'word' if m.group(1)[-1] == 'W' else 'dword')
    m = re.match(r'^%(DB\d+)\.DB([WDB])(\d+)$', s)
    if m:
        db = m.group(1)
        typ = m.group(2)
        off = int(m.group(3))
        sz = 'byte' if typ == 'B' else ('word' if typ == 'W' else 'dword')
        return (db, off, 0, sz)
    return (None, None, None, None)

def to_modbus(area, byte_off, bit_off, size_hint):
    if area is None:
        return ("", "", "")
    if area == 'I':
        addr = 10001 + byte_off * 8 + bit_off
        return (str(addr), "离散输入 (1xxxx)", "FC02")
    elif area == 'Q':
        addr = 1 + byte_off * 8 + bit_off
        return (str(addr).zfill(5), "线圈 (0xxxx)", "FC01/05/15")
    elif area == 'M':
        addr = 1 + byte_off * 8 + bit_off
        return (str(addr).zfill(5), "线圈 (0xxxx)-M区", "FC01/05/15")
    elif area == 'MW':
        addr = 40001 + byte_off
        return (str(addr), "保持寄存器 (4xxxx)", "FC03/06/16")
    elif area == 'MD':
        addr = 40001 + byte_off
        return (str(addr), "保持寄存器 (4xxxx)-双字", "FC03/06/16 (2 regs)")
    elif area == 'IW':
        addr = 30001 + byte_off
        return (str(addr), "输入寄存器 (3xxxx)", "FC04")
    elif area and area.startswith('DB'):
        off = byte_off // 2 if size_hint in ('word','dword') else byte_off
        return (str(40001 + off), f"保持寄存器 (4xxxx)-{area}", "FC03/06/16")
    return ("", "", "")


# ── 3. 提取变量名和地址 ──
tags = []

if sep:
    # 找到列索引
    name_idx = None
    addr_idx = None
    dtype_idx = None
    comment_idx = None
    for i, h in enumerate(headers):
        hlow = h.lower().strip()
        if hlow in ('name', '名称', 'tag name', '变量名', 'tagname'):
            name_idx = i
        elif hlow in ('address', '地址', 'logicaladdress', 'logical address'):
            addr_idx = i
        elif hlow in ('data type', 'datatype', '数据类型', 'type'):
            dtype_idx = i
        elif hlow in ('comment', '注释', 'description'):
            comment_idx = i

    # 如果没找到精确匹配，尝试模糊匹配
    if name_idx is None:
        for i, h in enumerate(headers):
            if '名' in h or 'name' in h.lower():
                name_idx = i
                break
    if addr_idx is None:
        for i, h in enumerate(headers):
            if '地址' in h or 'addr' in h.lower():
                addr_idx = i
                break
    if dtype_idx is None:
        for i, h in enumerate(headers):
            if '类型' in h or 'type' in h.lower() or 'data' in h.lower():
                dtype_idx = i
                break

    print(f"列映射: name={name_idx}, addr={addr_idx}, dtype={dtype_idx}, comment={comment_idx}")

    for line in data_lines:
        cols = line.split(sep)
        name = cols[name_idx].strip() if name_idx is not None and name_idx < len(cols) else ''
        addr = cols[addr_idx].strip() if addr_idx is not None and addr_idx < len(cols) else ''
        dtype = cols[dtype_idx].strip() if dtype_idx is not None and dtype_idx < len(cols) else ''
        comment = cols[comment_idx].strip() if comment_idx is not None and comment_idx < len(cols) else ''

        if not name:
            continue

        # 筛选 e1_ 变量
        if name.lower().startswith('e1_'):
            area, bo, bit, sz = parse_logical_address(addr)
            mb_addr, mb_area, fc = to_modbus(area, bo, bit, sz)
            tags.append({
                'Name': name, 'DataType': dtype, 'LogicalAddress': addr,
                'Comment': comment, 'ModbusAddress': mb_addr,
                'ModbusArea': mb_area, 'FunctionCode': fc
            })
else:
    # 纯名称列表，只有名称没有地址
    for line in data_lines:
        name = line.strip()
        if name.lower().startswith('e1_'):
            tags.append({
                'Name': name, 'DataType': '', 'LogicalAddress': '',
                'Comment': '', 'ModbusAddress': '', 'ModbusArea': '', 'FunctionCode': ''
            })

print(f"\n提取到 {len(tags)} 个 e1_ 变量")

# ── 4. 输出 Excel ──
if tags:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "e1_Modbus地址"
    
    xl_headers = ['变量名', '数据类型', '逻辑地址', 'Modbus地址', 'Modbus区域', '功能码', '注释']
    hf = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    
    for ci, h in enumerate(xl_headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hf
        cell.alignment = Alignment(horizontal='center')
    
    for ri, tag in enumerate(tags, 2):
        for ci, key in enumerate(['Name','DataType','LogicalAddress','ModbusAddress','ModbusArea','FunctionCode','Comment'], 1):
            ws.cell(row=ri, column=ci, value=tag.get(key, ''))
    
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{len(tags)+1}'
    
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OUTPUT_DIR / f"一号挤出机_e1_Modbus地址_{ts}.xlsx"
    wb.save(str(out))
    print(f"\n[✓] Excel 已保存: {out}")

# ── 5. 显示结果摘要 ──
if tags:
    print(f"\n=== 变量列表 ===")
    for t in tags[:10]:
        print(f"  {t['Name']:40s} {t['LogicalAddress']:12s} → {t['ModbusAddress']:6s} {t['ModbusArea']}")
    if len(tags) > 10:
        print(f"  ... 共 {len(tags)} 个")
    
    # 统计
    with_addr = sum(1 for t in tags if t['LogicalAddress'])
    without_addr = sum(1 for t in tags if not t['LogicalAddress'])
    print(f"\n有地址: {with_addr}, 无地址: {without_addr}")
