#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys,re,datetime,os
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter

SCR=Path(__file__).parent.resolve()
OUT=SCR/'导出结果';DOC=SCR/'设备说明'
OUT.mkdir(parents=True,exist_ok=True);DOC.mkdir(parents=True,exist_ok=True)

wb=openpyxl.load_workbook(SCR/'手动导出'/'PLCTags.xlsx');ws=wb['PLC Tags']
vars_raw=[]
for r in range(2,ws.max_row+1):
    n=str(ws.cell(row=r,column=1).value or '').strip()
    p=str(ws.cell(row=r,column=2).value or '').strip()
    d=str(ws.cell(row=r,column=3).value or '').strip()
    a=str(ws.cell(row=r,column=4).value or '').strip()
    c=str(ws.cell(row=r,column=5).value or '').strip()
    if n:vars_raw.append({'n':n,'p':p,'d':d,'a':a,'c':c})
wb.close()

def pa(s):
    if not s:return(None,None,None,None)
    s=s.strip()
    m=re.match(r'^%([IQM])(\d+)\.(\d+)$',s)
    if m:return(m.group(1),int(m.group(2)),int(m.group(3)),'bit')
    m=re.match(r'^%(MW|MD|IW|ID|QW)(\d+)$',s)
    if m:a=m.group(1);return(a,int(m.group(2)),0,'dword'if a in('MD','ID')else'word')
    m=re.match(r'^%(DB\d+)\.DB([WDB])(\d+)$',s)
    if m:db,typ,off=m.group(1),m.group(2),int(m.group(3));sz='byte'if typ=='B'else('word'if typ=='W'else'dword');return(db,off,0,sz)
    return(None,None,None,None)

def mb(area,bo,bit,sz):
    if area is None:return('','','')
    if area=='I':return(str(10001+bo*8+bit),'离散输入(1xxxx)','FC02')
    if area=='Q':return(str(1+bo*8+bit).zfill(5),'线圈(0xxxx)','FC01/05/15')
    if area=='M':return(str(1+bo*8+bit).zfill(5),'线圈(0xxxx)-M区','FC01/05/15')
    if area=='MW':return(str(40001+bo),'保持寄存器(4xxxx)','FC03/06/16')
    if area=='MD':return(str(40001+bo),'保持寄存器(4xxxx)-双字','FC03/06/16(2regs)')
    if area=='IW':return(str(30001+bo),'输入寄存器(3xxxx)','FC04')
    if area=='QW':return(str(40001+bo),'保持寄存器(4xxxx)-QW','FC03/06/16')
    if area and area.startswith('DB'):
        off=bo//2 if sz in('word','dword')else bo
        return(str(40001+off),f'保持寄存器(4xxxx)-{area}','FC03/06/16')
    return('','','')

def tp(v):
    nl=v['n'].lower()
    if nl.endswith('_di'):return'DI'
    if re.search(r'_do$|_do_spare$|_start_do$|_power_do$',nl):return'DO'
    if nl.endswith('_ai'):return'AI'
    if nl.endswith('_aq')or'speed_aq'in nl or'set_aq'in nl:return'AQ'
    return'OTHER'

STATIONS=[
{'id':'station_1','name':'一号挤出机','short':'1#extruder','prefix':'e1_','tt':['E1','默认变量表'],'pat':[],'file':'一号挤出机变量清单.md'},
{'id':'station_2','name':'辅机上料','short':'Feeder','prefix':'e0_','tt':['E0','Feeding_Puller'],'pat':['Material','Feeding'],'file':'辅机上料变量清单.md'},
{'id':'station_3','name':'二号挤出机','short':'2#extruder','prefix':'e3_','tt':['E3'],'pat':[],'file':'二号挤出机变量清单.md'},
{'id':'station_4','name':'测径检测','short':'Measurement','prefix':'','tt':['Sikora_1','Sikora_2','LaserMarker','Tunnel'],'pat':['Sikora','Laser','Tunnel','PN_AW','OD_'],'file':'测径检测变量清单.md'},
{'id':'station_5','name':'主牵引机1','short':'Puller1','prefix':'','tt':['Puller_1','Minor_Puller'],'pat':['Master_Puller_1','Puller_1','Minor_Puller'],'file':'主牵引机1变量清单.md'},
{'id':'station_6','name':'切割机','short':'Cutter','prefix':'','tt':['State'],'pat':['Cutter','Cut_'],'file':'切割机变量清单.md'},
{'id':'station_7','name':'主牵引机2','short':'Puller2','prefix':'','tt':['Puller_2'],'pat':['Master_Puller_2','Puller_2'],'file':'主牵引机2变量清单.md'},
{'id':'station_8','name':'收卷机','short':'Winder','prefix':'','tt':['Lucas_Knitting'],'pat':['Lucas','Knitting','Belt','Winder','收卷'],'file':'收卷机变量清单.md'},
]

for st in STATIONS:
    st['vars']=[];st['di']=[];st['do']=[];st['ai']=[];st['aq']=[];st['other']=[]

assigned=set()
for v in vars_raw:
    for st in STATIONS:
        nl=v['n'].lower();pl=v['p'].lower()
        ok=False
        if st['prefix'] and nl.startswith(st['prefix'].lower()):ok=True
        if not ok:
            for t in st['tt']:
                if t.lower()in pl:ok=True;break
        if not ok:
            for p in st['pat']:
                if p.lower()in nl or p.lower()in pl:ok=True;break
        if ok:
            st['vars'].append(v);assigned.add(v['n'])
            t=tp(v)
            if t=='DI':st['di'].append(v)
            elif t=='DO':st['do'].append(v)
            elif t=='AI':st['ai'].append(v)
            elif t=='AQ':st['aq'].append(v)
            else:st['other'].append(v)
            break

def mk_section(title,tags):
    if not tags:return''
    l=[f'## {title} ({len(tags)}个)','',
       '| 变量名 | 数据类型 | 逻辑地址 | Modbus地址 | 区域 | 功能码 | 注释 |',
       '|--------|---------|---------|-----------|------|------|------|']
    for t in sorted(tags,key=lambda x:x['n']):
        a,bo,b,sz=pa(t['a']);ma,ar,fc=mb(a,bo,b,sz)
        l.append(f"| `{t['n']}` | {t['d']} | `{t['a']}` | {ma} | {ar} | {fc} | {t['c']} |")
    l.append('');return'\n'.join(l)

area_colors={'I':'DAEEF3','Q':'D5F5E3','M':'FCF3CF','MW':'FADBD8','MD':'FADBD8','IW':'E8DAEF','ID':'E8DAEF','QW':'E8DAEF'}

for st in STATIONS:
    vs=st['vars']
    if not vs:print(f"[!] {st['name']}: 0变量");continue
    print(f"{st['name']}: {len(vs)}个 (DI:{len(st['di'])} DO:{len(st['do'])} AI:{len(st['ai'])} AQ:{len(st['aq'])} OTHER:{len(st['other'])})")
    
    # Excel
    wb2=openpyxl.Workbook();ws2=wb2.active;ws2.title=st['short'][:31]
    xh=['变量名','数据类型','逻辑地址','Modbus地址','Modbus区域','功能码','注释','变量表']
    hf=PatternFill(start_color='2F5496',end_color='2F5496',fill_type='solid')
    hfont=Font(bold=True,color='FFFFFF',size=11)
    for ci,h in enumerate(xh,1):c=ws2.cell(row=1,column=ci,value=h);c.font=hfont;c.fill=hf;c.alignment=Alignment(horizontal='center')
    for ri,v in enumerate(vs,2):
        a,bo,b,sz=pa(v['a']);ma,ar,fc=mb(a,bo,b,sz)
        for ci,val in enumerate([v['n'],v['d'],v['a'],ma,ar,fc,v['c'],v['p']],1):ws2.cell(row=ri,column=ci,value=val)
        if a in area_colors:
            fill=PatternFill(start_color=area_colors[a],end_color=area_colors[a],fill_type='solid')
            for ci in range(1,9):ws2.cell(row=ri,column=ci).fill=fill
    for ci,w in enumerate([32,14,18,16,28,18,30,20],1):ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.freeze_panes='A2';ws2.auto_filter.ref=f'A1:H{len(vs)+1}'
    ts=datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    xo=OUT/f"{st['short']}_Modbus_{ts}.xlsx";wb2.save(str(xo))
    
    # Markdown
    md=[]
    md.append(f"# {st['name']} ({st['short']}) 变量清单")
    md.append('');md.append(f"> 数据来源: TIA Portal 手动导出 PLCTags.xlsx")
    md.append(f"> 导出时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    md.append(f"> 变量总数: {len(vs)}");md.append('');md.append('---');md.append('')
    md.append('## Modbus 地址换算规则');md.append('')
    md.append('| 逻辑地址 | Modbus 区域 | 地址公式 | 功能码 |')
    md.append('|---------|-----------|---------|-------|')
    md.append('| `%Ix.y` | 离散输入(1xxxx) | `10001+x*8+y` | FC02 |')
    md.append('| `%Qx.y` | 线圈(0xxxx) | `1+x*8+y` | FC01/05/15 |')
    md.append('| `%Mx.y` | 线圈(0xxxx)-M区 | `1+x*8+y` | FC01/05/15 |')
    md.append('| `%MWx` | 保持寄存器(4xxxx) | `40001+x` | FC03/06/16 |')
    md.append('| `%IWx` | 输入寄存器(3xxxx) | `30001+x` | FC04 |')
    md.append('| `%QWx` | 保持寄存器(4xxxx)-QW | `40001+x` | FC03/06/16 |')
    md.append('')
    md.append('> 实际 Modbus 地址取决于网关设备（CoTrust CTH2-277PN / ADFweb HD67607）映射配置。')
    md.append('')
    for title,tags in [('一、DI 数字量输入',st['di']),('二、DO 数字量输出',st['do']),('三、AI 模拟量输入',st['ai']),('四、AQ 模拟量输出',st['aq']),('五、其他变量',st['other'])]:
        sec=mk_section(title,tags)
        if sec:md.append('---');md.append('');md.append(sec)
    mp=DOC/st['file']
    with open(mp,'w',encoding='utf-8')as f:f.write('\n'.join(md))
    print(f'  -> {xo.name} / {st["file"]}')

# 未分配
unas=[v for v in vars_raw if v['n'] not in assigned]
if unas:
    print(f'\n未分配({len(unas)}):')
    for v in unas[:15]:print(f'  {v["n"]:40s} Path={v["p"]}')

print('\n[✓] 完成!')
