#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成 TIA Portal 模拟变量表（测试用）
=====================================
在 手动导出/ 目录下创建模拟的 TIA Portal 变量表 Excel 文件，
用于测试 parse_exported_variables.py 的解析功能。

用法:
    python generate_test_data.py

输出:
    手动导出/PLC变量表_模拟.xlsx     ← PLC 变量（中文列名）
    手动导出/HMI变量_模拟.xlsx       ← HMI 变量（英文列名）
    手动导出/DB块_模拟.xlsx          ← DB 块变量

作者: CodeWhale
日期: 2026-06-30
"""

import sys
import datetime
from pathlib import Path

# Windows 控制台 GBK 编码兼容
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("需要 openpyxl。请运行: pip install openpyxl")
    exit(1)

OUTPUT_DIR = Path(__file__).parent / "手动导出"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def create_plc_tags_xlsx():
    """生成模拟 PLC 变量表（中文列名）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "默认变量表"

    # ── 表头（中文） ──
    headers = ["名称", "数据类型", "地址", "注释", "变量表", "PLC设备"]
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── 模拟数据 ──
    tags = [
        # (名称, 数据类型, 地址, 注释)
        ("Motor_Start",      "Bool",    "%I0.0",   "电机启动按钮"),
        ("Motor_Stop",       "Bool",    "%I0.1",   "电机停止按钮"),
        ("Emergency_Stop",   "Bool",    "%I0.2",   "急停按钮(NC)"),
        ("Proximity_Sensor", "Bool",    "%I0.3",   "接近传感器"),
        ("Auto_Mode",        "Bool",    "%I0.4",   "自动模式选择开关"),
        ("Speed_Setpoint",   "Int",     "%MW10",   "速度设定值 (0-1500 rpm)"),
        ("Temperature",      "Int",     "%MW12",   "当前温度 ×10"),
        ("Pressure",         "Real",    "%MD20",   "压力反馈 (bar)"),
        ("Length_Counter",   "DInt",    "%MD24",   "计米器 (mm)"),
        ("Alarm_Word",       "Word",    "%MW30",   "报警字"),
        ("Motor_Run",        "Bool",    "%Q0.0",   "电机运行输出"),
        ("Valve_Open",       "Bool",    "%Q0.1",   "阀门打开"),
        ("Heater_On",        "Bool",    "%Q0.2",   "加热器启动"),
        ("Indicator_Green",  "Bool",    "%Q0.3",   "绿色指示灯"),
        ("Indicator_Red",    "Bool",    "%Q0.4",   "红色指示灯(故障)"),
        ("Fault_Reset",      "Bool",    "%M0.0",   "故障复位"),
        ("Cycle_Complete",   "Bool",    "%M0.1",   "周期完成标志"),
        ("Production_Done",  "Bool",    "%M0.2",   "生产完成"),
        ("Step_Number",      "Int",     "%MW40",   "当前步骤号"),
        ("Recipe_Number",    "Int",     "%MW42",   "配方号"),
        ("Machine_State",    "Int",     "%MW44",   "机器状态 (0=停机,1=运行,2=暂停)"),
        ("Target_Length",    "DInt",    "%MD50",   "目标长度 (mm)"),
        ("Actual_Speed",     "Real",    "%MD54",   "实际转速 (rpm)"),
        ("Cycle_Time",       "Time",    "%MD60",   "周期时间"),
        ("Barcode_Data",     "String[32]", "%DB1.DBB0", "条码数据"),
        ("Par_Speed_Max",    "Int",     "%DB1.DBW32",  "参数:最大速度"),
        ("Par_Temp_Set",     "Int",     "%DB1.DBW34",  "参数:设定温度"),
        ("Par_Pressure_Min", "Real",    "%DB1.DBD36",  "参数:最小压力"),
        ("HMI_Start",        "Bool",    "%M10.0",  "HMI启动按钮"),
        ("HMI_Speed_Set",    "Int",     "%MW100",  "HMI速度设定"),
    ]

    for row_idx, (name, dtype, addr, comment) in enumerate(tags, 2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=dtype)
        ws.cell(row=row_idx, column=3, value=addr)
        ws.cell(row=row_idx, column=4, value=comment)
        ws.cell(row=row_idx, column=5, value="默认变量表")
        ws.cell(row=row_idx, column=6, value="S7-1200 station_1")

    # 列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20

    # 冻结首行
    ws.freeze_panes = "A2"

    # ── 第二个 sheet：系统变量 ──
    ws2 = wb.create_sheet("系统变量")
    sys_headers = ["名称", "数据类型", "地址", "注释"]
    for col_idx, h in enumerate(sys_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    sys_tags = [
        ("FirstScan",      "Bool",  "%M1.0",  "首次扫描标志"),
        ("AlwaysTRUE",     "Bool",  "%M1.1",  "常ON"),
        ("AlwaysFALSE",    "Bool",  "%M1.2",  "常OFF"),
        ("Clock_1Hz",      "Bool",  "%M1.3",  "1Hz时钟脉冲"),
        ("Clock_10Hz",     "Bool",  "%M1.4",  "10Hz时钟脉冲"),
        ("OB1_ScanTime",   "Int",   "%MW200", "OB1扫描时间(ms)"),
    ]
    for row_idx, (name, dtype, addr, comment) in enumerate(sys_tags, 2):
        ws2.cell(row=row_idx, column=1, value=name)
        ws2.cell(row=row_idx, column=2, value=dtype)
        ws2.cell(row=row_idx, column=3, value=addr)
        ws2.cell(row=row_idx, column=4, value=comment)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 25

    path = OUTPUT_DIR / "PLC变量表_模拟.xlsx"
    wb.save(str(path))
    print(f"[✓] {path.name} ({len(tags) + len(sys_tags)} 个变量, 2 sheets)")


def create_hmi_tags_xlsx():
    """生成模拟 HMI 变量表（英文列名）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HMI_Tags"

    headers = ["Name", "Data type", "Address", "Comment", "Connection"]
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    hmi_tags = [
        ("HMI_Start_PB",        "Bool",   "S7-1200 station_1/Motor_Start",     "启动按钮"),
        ("HMI_Stop_PB",         "Bool",   "S7-1200 station_1/Motor_Stop",      "停止按钮"),
        ("HMI_Speed_SP",        "Int",    "S7-1200 station_1/HMI_Speed_Set",   "速度设定输入"),
        ("HMI_Actual_Speed",    "Real",   "S7-1200 station_1/Actual_Speed",    "实际速度显示"),
        ("HMI_Temperature",     "Int",    "S7-1200 station_1/Temperature",     "温度显示"),
        ("HMI_Pressure",        "Real",   "S7-1200 station_1/Pressure",        "压力显示"),
        ("HMI_Length",          "DInt",   "S7-1200 station_1/Length_Counter",  "计米显示"),
        ("HMI_Step_Num",        "Int",    "S7-1200 station_1/Step_Number",     "步骤号显示"),
        ("HMI_Machine_State",   "Int",    "S7-1200 station_1/Machine_State",   "状态显示"),
        ("HMI_Recipe_Num",      "Int",    "S7-1200 station_1/Recipe_Number",   "配方号"),
        ("HMI_Alarm_Display",   "Word",   "S7-1200 station_1/Alarm_Word",     "报警显示"),
        ("HMI_Green_Lamp",      "Bool",   "S7-1200 station_1/Indicator_Green", "绿灯指示"),
        ("HMI_Red_Lamp",        "Bool",   "S7-1200 station_1/Indicator_Red",   "红灯指示"),
        ("HMI_Auto_Mode",       "Bool",   "S7-1200 station_1/Auto_Mode",       "自动模式指示"),
        ("HMI_Cycle_Complete",  "Bool",   "S7-1200 station_1/Cycle_Complete",  "周期完成指示"),
        ("HMI_Target_Len",      "DInt",   "S7-1200 station_1/Target_Length",   "目标长度显示"),
    ]

    for row_idx, (name, dtype, addr, comment) in enumerate(hmi_tags, 2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=dtype)
        ws.cell(row=row_idx, column=3, value=addr)
        ws.cell(row=row_idx, column=4, value=comment)
        ws.cell(row=row_idx, column=5, value="HMI_Connection_1")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A2"

    path = OUTPUT_DIR / "HMI变量_模拟.xlsx"
    wb.save(str(path))
    print(f"[✓] {path.name} ({len(hmi_tags)} 个 HMI 变量)")


def create_db_vars_xlsx():
    """生成模拟 DB 块变量表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DB_Recipe"

    headers = ["名称", "数据类型", "地址", "注释", "所属块"]
    header_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    db_vars = [
        ("Recipe_ID",         "Int",        "DB1.DBW0",  "配方ID"),
        ("Recipe_Name",       "String[16]", "DB1.DBB2",  "配方名称"),
        ("Speed_Set",         "Int",        "DB1.DBW18", "速度设定"),
        ("Temp_Set",          "Int",        "DB1.DBW20", "温度设定"),
        ("Pressure_Min",      "Real",       "DB1.DBD22", "最小压力"),
        ("Pressure_Max",      "Real",       "DB1.DBD26", "最大压力"),
        ("Length_Target",     "DInt",       "DB1.DBD30", "目标长度"),
        ("Tolerance_Plus",    "Real",       "DB1.DBD34", "正公差"),
        ("Tolerance_Minus",   "Real",       "DB1.DBD38", "负公差"),
        ("Dwell_Time",        "Time",       "DB1.DBD42", "保压时间"),
        ("Cooling_Time",      "Time",       "DB1.DBD46", "冷却时间"),
        ("Max_Current",       "Real",       "DB1.DBD50", "最大电流"),
    ]

    for row_idx, (name, dtype, addr, comment) in enumerate(db_vars, 2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=dtype)
        ws.cell(row=row_idx, column=3, value=addr)
        ws.cell(row=row_idx, column=4, value=comment)
        ws.cell(row=row_idx, column=5, value="DB1 (Recipe)")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.freeze_panes = "A2"

    # ── 第二个 sheet：DB_Machine ──
    ws2 = wb.create_sheet("DB_Machine_Params")
    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    db2_vars = [
        ("Motor_Max_RPM",     "Int",    "DB2.DBW0",  "电机最大转速"),
        ("Motor_Min_RPM",     "Int",    "DB2.DBW2",  "电机最小转速"),
        ("Accel_Ramp",        "Int",    "DB2.DBW4",  "加速斜坡(ms)"),
        ("Decel_Ramp",        "Int",    "DB2.DBW6",  "减速斜坡(ms)"),
        ("PID_Kp",            "Real",   "DB2.DBD8",  "PID比例系数"),
        ("PID_Ki",            "Real",   "DB2.DBD12", "PID积分系数"),
        ("PID_Kd",            "Real",   "DB2.DBD16", "PID微分系数"),
        ("Total_Runtime_H",   "DInt",   "DB2.DBD20", "累计运行时间(小时)"),
        ("Cycle_Count",       "DInt",   "DB2.DBD24", "生产周期计数"),
        ("Maintenance_Flag",  "Bool",   "DB2.DBX28.0", "维护提示标志"),
    ]
    for row_idx, (name, dtype, addr, comment) in enumerate(db2_vars, 2):
        ws2.cell(row=row_idx, column=1, value=name)
        ws2.cell(row=row_idx, column=2, value=dtype)
        ws2.cell(row=row_idx, column=3, value=addr)
        ws2.cell(row=row_idx, column=4, value=comment)
        ws2.cell(row=row_idx, column=5, value="DB2 (MachineParams)")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 28
    ws2.column_dimensions["E"].width = 22

    path = OUTPUT_DIR / "DB块_模拟.xlsx"
    wb.save(str(path))
    print(f"[✓] {path.name} ({len(db_vars) + len(db2_vars)} 个 DB 变量, 2 sheets)")


def main():
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"生成 TIA Portal 模拟变量表 ({stamp})")
    print(f"输出目录: {OUTPUT_DIR}\n")

    create_plc_tags_xlsx()
    create_hmi_tags_xlsx()
    create_db_vars_xlsx()

    print(f"\n[✓] 完成！共生成 3 个测试文件\n")
    print(f"现在可以运行解析脚本验证:")
    print(f"  python parse_exported_variables.py")
    print(f"  python parse_exported_variables.py \"{OUTPUT_DIR}\"")


if __name__ == "__main__":
    main()
