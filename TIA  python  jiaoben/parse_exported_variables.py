#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
TIA Portal 手动导出变量解析工具（备选方案）
============================================
适用场景：没有 TIA Portal Openness API 时，在 TIA Portal 中手动导出
         变量表为 .xlsx 文件，然后用此脚本读取、合并和分析。

在 TIA Portal 中手动导出的方法：
  1. 项目树 → PLC_1 → "PLC 变量" → 双击打开变量表
  2. 在上方表格中 Ctrl+A → Ctrl+C
  3. 粘贴到 Excel，保存到指定目录
  4. 同样操作导出 "HMI 变量" 和 "DB 块"

  或使用 TIA Portal 内置导出：
  右键变量表文件夹 → "导出" → 另存为 .xlsx

用法:
    python parse_exported_variables.py                    # 自动读取 手动导出/ 目录
    python parse_exported_variables.py <目录路径>          # 指定读取目录
    python parse_exported_variables.py --merge            # 合并为单一变量清单

输出:
    导出结果/变量汇总_<时间戳>.xlsx      # 统一格式的变量清单

作者: CodeWhale
日期: 2026-06-30
"""

import sys
import datetime
import re
from pathlib import Path

# Windows 控制台 GBK 编码兼容
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要 openpyxl 库。请运行: pip install openpyxl")
    sys.exit(1)


# ────────────────────────────────────────────────────────────
#  配置
# ────────────────────────────────────────────────────────────

BASE_DIR = Path(r"C:\Users\Administrator\Documents\Project\python  jiaoben")
INPUT_DIR = BASE_DIR / "手动导出"
OUTPUT_DIR = BASE_DIR / "导出结果"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR.mkdir(parents=True, exist_ok=True)


# ════════════════════════════════════════════════════════════
#  TIA Portal 导出格式识别
# ════════════════════════════════════════════════════════════

# TIA Portal 导出的常见列名变体（中英文）
HEADER_PATTERNS = {
    "Name":      ["名称", "Name", "变量", "Variable", "变量名", "Tag Name"],
    "DataType":  ["数据类型", "Data type", "数据类型", "Type", "类型"],
    "Address":   ["地址", "Address", "逻辑地址", "Logical Address", "I地址"],
    "Comment":   ["注释", "Comment", "说明", "Description"],
    "TagTable":  ["变量表", "Tag table", "所属表", "Table"],
    "PlcDevice": ["PLC设备", "PLC Device", "设备"],
    "DefaultValue": ["初始值", "Start value", "Default", "默认值"],
}


def identify_headers(headers):
    """将 TIA Portal 导出的中文列名映射到标准英文键"""
    mapping = {}
    for col_name in headers:
        col_lower = col_name.strip().lower()
        for std_key, patterns in HEADER_PATTERNS.items():
            for pat in patterns:
                if pat.lower() in col_lower:
                    mapping[col_name] = std_key
                    break
            if col_name in mapping:
                break
    return mapping


def read_tia_export(file_path):
    """
    读取 TIA Portal 导出的 Excel 文件，自动识别格式。

    返回: (list[dict], str)  — 变量列表 + 识别的类别（PLC Tags / HMI Tags / DB）
    """
    print(f"\n读取: {file_path.name}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"  [✗] 无法打开文件: {e}")
        return [], "unknown"

    all_vars = []
    sheet_type = "unknown"

    for ws in wb.worksheets:
        # 读取表头
        headers = [cell.value for cell in ws[1]]
        if not headers or all(h is None for h in headers):
            continue

        # 过滤空列
        valid_cols = [(i, h) for i, h in enumerate(headers) if h is not None]
        if len(valid_cols) < 2:
            continue

        # 识别列名映射
        col_map = identify_headers([h for _, h in valid_cols])
        if not col_map:
            print(f"  [!] Sheet '{ws.title}' 无法识别格式，跳过")
            continue

        # 判断类型
        if "Connection" in [str(c) for c in headers if c] or "HMI" in ws.title.upper():
            sheet_type = "HMI Tags"
        elif "DB" in ws.title.upper() or "DataBlock" in ws.title:
            sheet_type = "DB Variables"
        elif "Tag" in ws.title or "变量" in ws.title or "tag" in ws.title.lower():
            sheet_type = "PLC Tags"
        else:
            sheet_type = "PLC Tags"

        print(f"  Sheet: {ws.title}  →  {sheet_type}  ({len(valid_cols)} 列)")

        # 读取数据行
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            var_data = {"SourceFile": file_path.name, "SheetType": sheet_type}
            empty = True

            for col_idx, (orig_idx, header_name) in enumerate(valid_cols):
                cell_value = row[col_idx] if col_idx < len(row) else None
                if cell_value is not None:
                    empty = False

                std_key = col_map.get(header_name)
                if std_key:
                    var_data[std_key] = cell_value

            if empty:
                continue

            # 确保有必需的键
            var_data.setdefault("Name", "")
            var_data.setdefault("DataType", "")
            var_data.setdefault("Address", "")
            var_data.setdefault("Comment", "")
            var_data.setdefault("TagTable", ws.title)
            var_data.setdefault("PlcDevice", "")

            all_vars.append(var_data)
            count += 1

        print(f"    共 {count} 行数据")

    return all_vars, sheet_type


# ════════════════════════════════════════════════════════════
#  变量分类与增强
# ════════════════════════════════════════════════════════════

def classify_address(address):
    """
    根据西门子 PLC 地址判断变量类型。

    返回: (类型, 区域, 偏移量)
    """
    address = str(address).strip()
    if not address:
        return "未分类", "", ""

    # 模式匹配（支持多种格式）
    patterns = [
        (r"^%I([\d.]+)",   "Input",   "I"),
        (r"^I([\d.]+)",    "Input",   "I"),
        (r"^%Q([\d.]+)",   "Output",  "Q"),
        (r"^Q([\d.]+)",    "Output",  "Q"),
        (r"^%M([\d.]+)",   "M",       "M"),
        (r"^M([\d.]+)",    "M",       "M"),
        (r"^%DB(\d+)\.(.*)", "DB",    "DB"),
        (r"^DB(\d+)\s*\.\s*(.*)", "DB", "DB"),
        (r"^\"(.*)\"",     "符号访问", ""),
        (r"^P#.*",         "指针",    ""),
    ]

    for pattern, category, area in patterns:
        m = re.match(pattern, address, re.IGNORECASE)
        if m:
            return category, area, m.group(1)

    return "其他", "", ""


def enhance_variables(variables):
    """增强变量信息：添加分类、地址解析、数据类型归类"""
    for var in variables:
        addr = var.get("Address", "")

        # 分类
        cat, area, offset = classify_address(addr)
        var["Category"] = cat
        var["Area"] = area
        var["Offset"] = offset

        # 数据类型归类
        dtype = str(var.get("DataType", "")).upper().strip()
        if dtype in ("BOOL", "BYTE", "WORD", "DWORD", "LWORD", "INT", "DINT", "LINT",
                       "SINT", "USINT", "UINT", "UDINT", "ULINT", "REAL", "LREAL",
                       "S5TIME", "TIME", "DATE", "TIME_OF_DAY", "TOD", "DT", "DTL",
                       "STRING", "WSTRING", "CHAR", "WCHAR"):
            var["TypeGroup"] = "基本类型"
        elif dtype.startswith("ARRAY"):
            var["TypeGroup"] = "数组"
        elif dtype in ("STRUCT", "UDT") or "STRUCT" in dtype:
            var["TypeGroup"] = "结构体"
        elif "DB_" in dtype or dtype.startswith("IEC_"):
            var["TypeGroup"] = "系统类型"
        elif dtype in ("COUNTER", "CTU", "CTD", "CTUD", "TIMER", "TON", "TOF", "TP", "TONR"):
            var["TypeGroup"] = "计数器/定时器"
        else:
            var["TypeGroup"] = "自定义类型"


# ════════════════════════════════════════════════════════════
#  合并与统计分析
# ════════════════════════════════════════════════════════════

def merge_variables(all_sources):
    """
    合并多个来源的变量列表，去重。

    去重依据: Name + Address + DataType
    """
    seen = set()
    merged = []
    duplicates = 0

    for var_list, src_type in all_sources:
        for var in var_list:
            key = (
                str(var.get("Name", "")).strip(),
                str(var.get("Address", "")).strip(),
                str(var.get("DataType", "")).strip(),
            )
            if key in seen:
                duplicates += 1
                continue
            seen.add(key)
            merged.append(var)

    if duplicates:
        print(f"\n  去重: 移除 {duplicates} 个重复变量")

    return merged


def generate_statistics(variables):
    """生成变量统计信息"""
    stats = {}

    # 按分类统计
    stats["by_category"] = {}
    for var in variables:
        cat = var.get("Category", "未分类")
        stats["by_category"][cat] = stats["by_category"].get(cat, 0) + 1

    # 按数据类型归类统计
    stats["by_type_group"] = {}
    for var in variables:
        tg = var.get("TypeGroup", "未知")
        stats["by_type_group"][tg] = stats["by_type_group"].get(tg, 0) + 1

    # 按源文件统计
    stats["by_source"] = {}
    for var in variables:
        src = var.get("SourceFile", "未知")
        stats["by_source"][src] = stats["by_source"].get(src, 0) + 1

    return stats


def print_statistics(stats, variables):
    """打印统计信息到控制台"""
    print("\n" + "=" * 60)
    print("  变量统计")
    print("=" * 60)
    print(f"  变量总数: {len(variables)}")

    print("\n  按地址区域:")
    for cat, count in sorted(stats["by_category"].items(), key=lambda x: -x[1]):
        bar = "█" * min(count // 5, 40)
        print(f"    {cat:12s} {count:5d}  {bar}")

    print("\n  按数据类型:")
    for tg, count in sorted(stats["by_type_group"].items(), key=lambda x: -x[1]):
        print(f"    {tg:16s} {count:5d}")

    print("\n  按来源文件:")
    for src, count in sorted(stats["by_source"].items(), key=lambda x: -x[1]):
        print(f"    {src[:60]}: {count}")


# ════════════════════════════════════════════════════════════
#  输出
# ════════════════════════════════════════════════════════════

def write_output(variables, output_path, stats=None):
    """将所有变量写入 Excel，格式化输出"""

    wb = openpyxl.Workbook()

    # ── Sheet 1: 全部变量 ──
    ws = wb.active
    ws.title = "全部变量"

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["SourceFile", "SheetType", "Name", "DataType", "TypeGroup",
               "Address", "Category", "Area", "Offset", "Comment", "PlcDevice",
               "TagTable", "DefaultValue"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 分类着色
    cat_colors = {
        "Input":     "DAEEF3",  # 浅蓝
        "Output":    "D5F5E3",  # 浅绿
        "M":         "FCF3CF",  # 浅黄
        "DB":        "FADBD8",  # 浅红
        "未分类":    "E8E8E8",  # 灰
    }

    for row_idx, var in enumerate(variables, 2):
        cat = var.get("Category", "未分类")
        fill = PatternFill(start_color=cat_colors.get(cat, "FFFFFF"),
                           end_color=cat_colors.get(cat, "FFFFFF"),
                           fill_type="solid")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=var.get(h, ""))
            cell.fill = fill
            cell.border = thin_border

    # 自动宽度
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(variables) + 1}"

    # ── Sheet 2: 按分类分组 ──
    if stats and variables:
        ws2 = wb.create_sheet("按分类分组")
        ws2.append(["分类", "数量", "占比"])
        for cat, count in sorted(stats["by_category"].items(), key=lambda x: -x[1]):
            pct = f"{count / len(variables) * 100:.1f}%"
            ws2.append([cat, count, pct])
        # 格式
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 10

    # ── Sheet 3: 按数据类型 ──
    if stats and variables:
        ws3 = wb.create_sheet("按数据类型")
        ws3.append(["数据类型归类", "数量", "占比"])
        for tg, count in sorted(stats["by_type_group"].items(), key=lambda x: -x[1]):
            pct = f"{count / len(variables) * 100:.1f}%"
            ws3.append([tg, count, pct])
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 10
        ws3.column_dimensions["C"].width = 10

    wb.save(str(output_path))
    print(f"\n[✓] 输出: {output_path}")


# ════════════════════════════════════════════════════════════
#  主函数
# ════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  TIA Portal 手动导出变量解析工具")
    print("  操作模式: 只读 — 仅解析已导出的文件")
    print("=" * 60)

    # 1. 确定输入目录
    if len(sys.argv) > 1 and not sys.argv[1].startswith("--"):
        input_dir = Path(sys.argv[1])
    else:
        input_dir = INPUT_DIR

    if not input_dir.exists():
        print(f"\n[✗] 输入目录不存在: {input_dir}")
        print(f"\n请先将 TIA Portal 导出的变量 Excel 文件放入此目录:")
        print(f"  {input_dir}")
        print(f"\n在 TIA Portal 中导出变量的方法:")
        print(f"  1. 打开 PLC 变量表 / HMI 变量表")
        print(f"  2. Ctrl+A 全选 → Ctrl+C 复制")
        print(f"  3. 粘贴到 Excel → 另存为 .xlsx → 放到上述目录")
        sys.exit(1)

    # 2. 读取所有 Excel 文件
    xlsx_files = sorted(input_dir.glob("*.xlsx"))
    if not xlsx_files:
        xlsx_files = sorted(input_dir.glob("*.xls"))

    if not xlsx_files:
        print(f"\n[!] 目录中没有 Excel 文件: {input_dir}")
        print("请将 TIA Portal 导出的 .xlsx 文件放入此目录后重试。")
        sys.exit(1)

    print(f"\n找到 {len(xlsx_files)} 个 Excel 文件:")
    for f in xlsx_files:
        print(f"  - {f.name}")

    # 3. 解析所有文件
    print("\n" + "=" * 60)
    print("  解析文件...")
    print("=" * 60)
    all_sources = []
    for f in xlsx_files:
        vars_list, src_type = read_tia_export(f)
        if vars_list:
            all_sources.append((vars_list, src_type))

    if not all_sources:
        print("\n[✗] 没有从文件中读取到任何变量数据。")
        sys.exit(1)

    # 4. 合并去重
    print("\n合并变量...")
    all_vars = merge_variables(all_sources)

    # 5. 增强变量信息
    print("增强变量信息...")
    enhance_variables(all_vars)

    # 6. 统计
    stats = generate_statistics(all_vars)
    print_statistics(stats, all_vars)

    # 7. 输出
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"变量汇总_{timestamp}.xlsx"
    write_output(all_vars, output_path, stats)

    # 8. 汇总
    print("\n" + "=" * 60)
    print(f"  完成！共处理 {len(xlsx_files)} 个文件，")
    print(f"  提取 {len(all_vars)} 个唯一定义变量")
    print(f"  输出文件: {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
