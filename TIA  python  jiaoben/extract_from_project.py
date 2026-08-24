#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""完整报告：从TIA Portal项目文件提取设备信息生成Excel"""
import sys
if sys.platform == 'win32':
    try: sys.stdout.reconfigure(encoding='utf-8')
    except: pass

import sqlite3, re, os, datetime
from pathlib import Path
from collections import Counter

PROJECT_ROOT = Path(os.environ['USERPROFILE']) / 'Documents' / 'Project' / 'JSB-25-081B（瑞源橡塑）'
OUTPUT_DIR = Path(r"C:\Users\Administrator\Documents\Project\python  jiaoben\导出结果")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HFILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HFONT = Font(bold=True, color="FFFFFF", size=11)

def collect_all_data():
    data = []

    # --- TPV 项目 GSD 设备 ---
    gsd_dir = PROJECT_ROOT / "JSB-25--081B(瑞源橡塑）TPV包纱管1.0" / "AdditionalFiles" / "GSD"
    for gsd_file in sorted(gsd_dir.glob('*')):
        if gsd_file.suffix.lower() not in ('.xml', '.gsdml', '.gsd'):
            continue
        try:
            with open(gsd_file, 'r', encoding='utf-8', errors='ignore') as fh:
                text = fh.read()
        except:
            continue

        vendor = re.search(r'VendorName[^>]*Value="([^"]+)"', text)
        vendor = vendor.group(1) if vendor else '?'
        family_m = re.search(r'Family[^>]*MainFamily="([^"]+)"[^>]*ProductFamily="([^"]+)"', text)
        family = f'{family_m.group(1)} / {family_m.group(2)}' if family_m else '?'
        mod_count = len(re.findall(r'<ModuleItem[^>]*ID=', text))
        vsm_count = len(re.findall(r'<VirtualSubmoduleItem[^>]*ID=', text))

        data.append({
            'Project': 'TPV包纱管',
            'Category': 'GSD设备',
            'File': gsd_file.name,
            'Vendor': vendor,
            'Description': family,
            'Detail1': f'模块: {mod_count}',
            'Detail2': f'虚拟子模块: {vsm_count}',
        })

    # --- TPV XRef ---
    xref = PROJECT_ROOT / "JSB-25--081B(瑞源橡塑）TPV包纱管1.0" / "XRef" / "XRef.db"
    if xref.exists():
        conn = sqlite3.connect(str(xref))
        cur = conn.cursor()
        for table in ['objs','rels','addrs','parts']:
            try:
                cur.execute(f"SELECT COUNT(*) FROM {table}")
                c = cur.fetchone()[0]
                data.append({'Project':'TPV包纱管','Category':'XRef统计','File':'XRef.db',
                             'Vendor':'','Description':table,'Detail1':f'{c} 行','Detail2':''})
            except: pass
        conn.close()

    # --- 针织机 GSD ---
    gsd_dir2 = PROJECT_ROOT / "针织机_V19（1214C+KTP700）" / "AdditionalFiles" / "GSD"
    for gsd_file in sorted(gsd_dir2.glob('*')):
        if gsd_file.suffix.lower() not in ('.xml', '.gsdml'):
            continue
        try:
            with open(gsd_file, 'r', encoding='utf-8', errors='ignore') as fh:
                text = fh.read()
        except: continue
        vendor = re.search(r'VendorName[^>]*Value="([^"]+)"', text)
        vendor = vendor.group(1) if vendor else 'Siemens'
        family_m = re.search(r'Family[^>]*MainFamily="([^"]+)"[^>]*ProductFamily="([^"]+)"', text)
        family = f'{family_m.group(1)} / {family_m.group(2)}' if family_m else 'Drives'
        data.append({'Project':'针织机','Category':'GSD设备','File':gsd_file.name,
                     'Vendor':vendor,'Description':family,'Detail1':'SINAMICS G120','Detail2':''})

    # --- 针织机 升级日志 ---
    log = PROJECT_ROOT / "针织机_V19（1214C+KTP700）" / "Logs" / "ConversionLog_16.0.0.0_to_19.0.0.0.xml"
    if log.exists():
        try:
            with open(log, 'r', encoding='utf-8') as fh:
                ltext = fh.read()
            devs = set(re.findall(r"DeviceDataObject '([^']+)'", ltext))
            for d in devs:
                data.append({'Project':'针织机','Category':'硬件设备','File':'升级日志',
                             'Vendor':'Siemens','Description':d,'Detail1':'','Detail2':''})
        except: pass

    # --- 针织机 XRef ---
    xref2 = PROJECT_ROOT / "针织机_V19（1214C+KTP700）" / "XRef" / "XRef.db"
    if xref2.exists():
        conn = sqlite3.connect(str(xref2))
        cur = conn.cursor()
        for table in ['objs','rels','addrs','parts']:
            try:
                cur.execute(f"SELECT COUNT(*) FROM {table}")
                c = cur.fetchone()[0]
                data.append({'Project':'针织机','Category':'XRef统计','File':'XRef.db',
                             'Vendor':'','Description':table,'Detail1':f'{c} 行','Detail2':''})
            except: pass
        conn.close()

    return data


def write_excel(data):
    wb = openpyxl.Workbook()
    headers = ['Project','Category','File','Vendor','Description','Detail1','Detail2']

    # Sheet 1: 全部
    ws = wb.active
    ws.title = "全部信息"
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HFONT; cell.fill = HFILL; cell.alignment = Alignment(horizontal='center')

    cat_colors = {
        'GSD设备': 'DAEEF3', 'XRef统计': 'FCF3CF',
        '硬件设备': 'D5F5E3', 'IO数据': 'FADBD8'
    }
    for ri, d in enumerate(data, 2):
        fill_c = cat_colors.get(d['Category'], 'FFFFFF')
        fill = PatternFill(start_color=fill_c, end_color=fill_c, fill_type="solid")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=ri, column=ci, value=d.get(h,''))
            cell.fill = fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{len(data)+1}'
    for i in range(1,8):
        ws.column_dimensions[get_column_letter(i)].width = [12,12,45,22,30,18,18][i-1]

    # Sheet 2: 按分类汇总
    ws2 = wb.create_sheet("分类汇总")
    ws2.append(['分类','数量'])
    for c in ws2[1]: c.font = HFONT; c.fill = HFILL
    cats = Counter(d['Category'] for d in data)
    for cat, count in cats.most_common():
        ws2.append([cat, count])
    ws2.column_dimensions['A'].width = 20

    out = OUTPUT_DIR / "项目设备变量汇总.xlsx"
    wb.save(str(out))
    print(f"[OK] {out}")
    return out


data = collect_all_data()
print(f"Collected {len(data)} items")
out = write_excel(data)

# Summary
cats = Counter(d['Category'] for d in data)
print(f"\nSummary:")
for cat, count in cats.most_common():
    print(f"  {cat}: {count}")
