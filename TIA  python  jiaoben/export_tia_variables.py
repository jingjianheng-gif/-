#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
TIA Portal 变量批量导出工具 — 基于 Siemens Openness API
================================================================
功能：读取运行中 TIA Portal 的项目，导出所有 PLC 变量、DB 块结构、
      HMI 变量、设备 I/O 映射到 Excel 文件。

前置条件：
  1. TIA Portal 已安装 Openness 组件（Professional / Advanced 版本）
  2. TIA Portal 正在运行，且所需项目已打开
  3. pythonnet 已安装：pip install pythonnet
  4. openpyxl 已安装：pip install openpyxl

作者: CodeWhale
日期: 2026-06-30
警告: 脚本仅执行读取操作，不修改任何项目数据。
"""

import os
import sys
import datetime
import traceback
from pathlib import Path

# ────────────────────────────────────────────────────────────
#  配置区域
# ────────────────────────────────────────────────────────────

# TIA Portal 版本配置 —— 根据实际安装路径修改
TIA_VERSIONS = {
    "V17": {
        "api_path": r"C:\Program Files\Siemens\Automation\Portal V17\PublicAPI",
        "assembly": "Siemens.Engineering",
    },
    "V19": {
        "api_path": r"C:\Program Files\Siemens\Automation\Portal V19\PublicAPI",
        "assembly": "Siemens.Engineering",
    },
}

# 项目名称关键字（用于匹配打开的项目，可不填即自动获取当前项目）
PROJECT_NAME_TPV = "TPV包纱管"
PROJECT_NAME_KNIT = "针织机"

# 输出目录
OUTPUT_DIR = Path(r"C:\Users\Administrator\Documents\Project\python  jiaoben\导出结果")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ════════════════════════════════════════════════════════════
#  依赖检查
# ════════════════════════════════════════════════════════════

def check_dependencies():
    """检查必要的 Python 包和 TIA Portal API 路径"""
    issues = []

    # 检查 pythonnet
    try:
        import clr
    except ImportError:
        issues.append(
            "pythonnet 未安装。请运行: pip install pythonnet\n"
            "注意: 64位 Python 需要 64位 pythonnet，确保 Python 架构与 TIA Portal 一致。"
        )

    # 检查 openpyxl
    try:
        import openpyxl
    except ImportError:
        issues.append("openpyxl 未安装。请运行: pip install openpyxl")

    # 检查 TIA Portal API 路径
    found_api = None
    for version_key, cfg in TIA_VERSIONS.items():
        api_path = Path(cfg["api_path"])
        dll_path = api_path / f"{cfg['assembly']}.dll"
        if dll_path.exists():
            found_api = (version_key, str(api_path))
            break

    if found_api is None:
        searched = "\n  ".join(v["api_path"] for v in TIA_VERSIONS.values())
        issues.append(
            f"未找到 TIA Portal Openness API DLL。\n"
            f"已搜索路径:\n  {searched}\n"
            f"请确认 TIA Portal 是否安装了 Openness 组件。\n"
            f"可在 TIA Portal 安装光盘中安装 'Openness' 选项。"
        )
    else:
        print(f"[✓] 找到 TIA Portal {found_api[0]} Openness API: {found_api[1]}")

    if issues:
        print("\n[✗] 依赖检查失败:")
        for i, issue in enumerate(issues, 1):
            print(f"\n  {i}. {issue}")
        return None

    print("[✓] 所有依赖检查通过")
    return found_api


# ════════════════════════════════════════════════════════════
#  环境设置与 TIA Portal 连接
# ════════════════════════════════════════════════════════════

def setup_environment(api_info):
    """配置 .NET 运行时环境并加载 Siemens.Engineering"""
    version_key, api_path = api_info

    # 将 API 目录加入 Python 搜索路径
    if api_path not in sys.path:
        sys.path.insert(0, api_path)

    import clr

    # 加载必要的 Siemens 程序集
    dlls = [
        f"{api_path}\\Siemens.Engineering.dll",
        f"{api_path}\\Siemens.Engineering.Hmi.dll",
        f"{api_path}\\Siemens.Engineering.HW.dll",
    ]
    for dll in dlls:
        if Path(dll).exists():
            clr.AddReference(dll)
            print(f"  [✓] 加载: {Path(dll).name}")
        else:
            print(f"  [!] 未找到: {Path(dll).name}（如果不需要可忽略）")

    # 导入命名空间
    from Siemens.Engineering import (
        TiaPortal,
        TiaPortalMode,
        Project,
        ProjectComposition,
    )
    from Siemens.Engineering.SW import (
        PlcSoftware,
        PlcBlock,
        PlcBlockGroup,
        PlcTagTable,
        PlcTagTableGroup,
    )
    from Siemens.Engineering.SW.Tags import PlcTag, PlcTagGroup
    from Siemens.Engineering.HW import Device, DeviceItem

    print(f"[✓] Siemens.Engineering 命名空间加载完毕\n")
    return Siemens, TiaPortal, TiaPortalMode


def connect_to_tia(Siemens, TiaPortal, TiaPortalMode):
    """连接到正在运行的 TIA Portal 进程"""
    # 方法一：连接到现有的 TIA Portal 进程
    print("正在连接到运行中的 TIA Portal 进程...")
    try:
        # TiaPortalMode.WithoutUI: 不显示额外的 UI（连接现有进程）
        # TiaPortalMode.WithUI: 显示 TIA Portal 界面（如果还没启动则启动）
        tia_portal = TiaPortal(TiaPortalMode.WithUI)
        print("[✓] 已连接到 TIA Portal")
        return tia_portal
    except Exception as e:
        print(f"[!] 连接失败: {e}")
        print("请确保 TIA Portal 已启动且项目已打开。尝试无 UI 模式...")
        try:
            tia_portal = TiaPortal(TiaPortalMode.WithoutUI)
            print("[✓] 已连接到 TIA Portal（无 UI 模式）")
            return tia_portal
        except Exception as e2:
            print(f"[✗] 无法连接到 TIA Portal: {e2}")
            sys.exit(1)


def get_open_project(tia_portal, project_filter=None):
    """获取当前打开的（或指定名称的）项目"""
    processes = list(tia_portal.Projects)
    if not processes:
        print("[✗] 没有打开的项目。请在 TIA Portal 中打开一个项目后重试。")
        sys.exit(1)

    print(f"当前有 {len(processes)} 个项目进程:")
    for i, proj in enumerate(processes):
        try:
            proj_path = proj.Path
            proj_name = Path(proj_path).name if proj_path else "(未知)"
            print(f"  {i+1}. {proj_name}")
            if project_filter and project_filter in proj_name:
                print(f"    → 匹配过滤条件，选择此项目")
                return proj
        except Exception:
            print(f"  {i+1}. (无法读取项目路径)")

    # 如果有过滤条件没匹配到，返回第一个
    if project_filter:
        print(f"  [!] 未找到匹配 '{project_filter}' 的项目，使用第一个项目")
    return processes[0]


# ════════════════════════════════════════════════════════════
#  变量导出函数
# ════════════════════════════════════════════════════════════

def export_plc_tags(project):
    """
    导出所有 PLC 变量表。

    返回: list[dict]  包含 Name, DataType, Address, Comment, TagTable, Device
    """
    all_tags = []

    for device in project.Devices:
        try:
            device_name = device.Name
        except Exception:
            device_name = "(未知设备)"

        for device_item in device.DeviceItems:
            try:
                item_name = device_item.Name
            except Exception:
                continue

            # 获取 PLC 软件
            try:
                sw_container = device_item.GetService("SoftwareContainer")
                if sw_container is None:
                    continue
                plc_software = sw_container.Software
                if plc_software is None:
                    continue
            except Exception:
                continue

            print(f"\n--- PLC: {item_name} ---")

            # 遍历变量表组
            try:
                tag_table_group = plc_software.TagTableGroup
            except Exception:
                print(f"  [!] 无法访问变量表组")
                continue

            _traverse_tag_tables(tag_table_group, all_tags, device_name, item_name)

    return all_tags


def _traverse_tag_tables(group, all_tags, device_name, plc_name, depth=0):
    """递归遍历变量表和变量表组"""
    prefix = "  " * depth

    # 遍历子组
    try:
        for subgroup in group.Groups:
            group_name = subgroup.Name
            print(f"{prefix}[变量表组] {group_name}")
            _traverse_tag_tables(
                subgroup, all_tags, device_name, plc_name, depth + 1
            )
    except Exception:
        pass

    # 遍历变量表
    try:
        for tag_table in group.TagTables:
            table_name = tag_table.Name
            tag_count = 0
            try:
                for tag in tag_table.Tags:
                    tag_info = _extract_tag_info(tag, table_name, device_name, plc_name)
                    all_tags.append(tag_info)
                    tag_count += 1
            except Exception as e:
                print(f"{prefix}  [!] 读取变量表 '{table_name}' 出错: {e}")
                continue
            print(f"{prefix}[变量表] {table_name} → {tag_count} 个变量")
    except Exception:
        pass


def _extract_tag_info(tag, table_name, device_name, plc_name):
    """从 PlcTag 对象提取变量信息"""
    return {
        "Device": device_name,
        "PLC": plc_name,
        "VariableTable": table_name,
        "Name": _safe_get(tag, "Name"),
        "DataType": _safe_get(tag, "DataType"),
        "LogicalAddress": _safe_get(tag, "LogicalAddress"),
        "Comment": _safe_get(tag, "Comment"),
    }


def export_db_blocks(project):
    """
    导出所有 DB 块中的变量结构。

    返回: list[dict]  包含 BlockName, BlockNumber, VariableName, DataType, Offset, Comment
    """
    all_db_vars = []

    for device in project.Devices:
        try:
            device_name = device.Name
        except Exception:
            device_name = "(未知)"

        for device_item in device.DeviceItems:
            try:
                sw_container = device_item.GetService("SoftwareContainer")
                if sw_container is None:
                    continue
                plc_software = sw_container.Software
                if plc_software is None:
                    continue
            except Exception:
                continue

            plc_name = device_item.Name
            print(f"\n--- DB 块: {plc_name} ---")

            try:
                block_group = plc_software.BlockGroup
            except Exception:
                print("  [!] 无法访问块组")
                continue

            _traverse_db_blocks(block_group, all_db_vars, device_name, plc_name)

    return all_db_vars


def _traverse_db_blocks(group, all_db_vars, device_name, plc_name, depth=0):
    """递归遍历块组，提取 DB 块变量结构"""
    prefix = "  " * depth

    try:
        for subgroup in group.Groups:
            _traverse_db_blocks(
                subgroup, all_db_vars, device_name, plc_name, depth + 1
            )
    except Exception:
        pass

    try:
        for block in group.Blocks:
            block_name = _safe_get(block, "Name", "")
            block_number = _safe_get(block, "Number", 0)
            block_type = _safe_get(block, "Type", "")

            # 只处理 DB 块
            if "DB" not in block_type and "DataBlock" not in str(type(block)):
                continue

            # 获取编程语言
            try:
                block_language = str(block.ProgrammingLanguage)
            except Exception:
                block_language = "(未知)"

            # 如果是 SCL/STL/LAD 类型，尝试打开并解析变量
            print(f"{prefix}[DB] {block_name} (编号 {block_number}, 语言 {block_language})")
            db_vars = _extract_db_variables(block, block_name, block_number, device_name, plc_name)
            all_db_vars.extend(db_vars)
    except Exception as e:
        print(f"{prefix}  [!] 遍历块时出错: {e}")


def _extract_db_variables(block, block_name, block_number, device_name, plc_name):
    """尝试从 DB 块中提取变量定义"""
    db_vars = []
    try:
        # 尝试获取块的接口定义
        block_interface = block.Interface
        if block_interface is None:
            return db_vars

        # 静态变量区域通常包含 DB 的数据结构
        sections = ["Static", "Input", "Output", "InOut", "Temp", "Constant"]
        for section_name in sections:
            try:
                section = getattr(block_interface, section_name, None)
                if section is None:
                    continue
                for member in section.Members:
                    db_vars.append({
                        "Device": device_name,
                        "PLC": plc_name,
                        "BlockName": block_name,
                        "BlockNumber": block_number,
                        "Section": section_name,
                        "VariableName": _safe_get(member, "Name"),
                        "DataType": _safe_get(member, "DataType"),
                        "Comment": _safe_get(member, "Comment"),
                    })
            except Exception:
                continue
    except Exception:
        pass
    return db_vars


def export_hmi_tags(project):
    """
    导出所有 HMI 变量。

    返回: list[dict]  包含 HmiDevice, TagName, Connection, DataType, Address, Comment
    """
    all_hmi_tags = []

    for device in project.Devices:
        try:
            device_name = device.Name
        except Exception:
            device_name = "(未知)"

        for device_item in device.DeviceItems:
            hmi_tags = _extract_hmi_tags_from_item(device_item, device_name)
            all_hmi_tags.extend(hmi_tags)

    return all_hmi_tags


def _extract_hmi_tags_from_item(device_item, device_name):
    """从单个 HMI 设备项中提取变量"""
    hmi_tags = []

    try:
        # 尝试获取 HMI 软件容器
        sw_container = device_item.GetService("SoftwareContainer")
        if sw_container is None or sw_container.Software is None:
            return hmi_tags

        hmi_software = sw_container.Software

        # 尝试获取 HMI 变量提供者
        try:
            tag_provider = hmi_software.GetService("HmiTagProvider")
            if tag_provider is not None:
                hmi_tags = _collect_hmi_tags(tag_provider, device_name, device_item.Name)
        except Exception:
            pass

        # 备用：尝试直接通过属性获取
        if not hmi_tags:
            try:
                tag_table = hmi_software.TagTable
                if tag_table is not None:
                    hmi_tags = _collect_hmi_tags_from_table(tag_table, device_name, device_item.Name)
            except Exception:
                pass

    except Exception:
        pass

    return hmi_tags


def _collect_hmi_tags(tag_provider, device_name, hmi_name):
    """通过 HmiTagProvider 收集变量"""
    tags = []
    try:
        for tag in tag_provider.Tags:
            tags.append({
                "HmiDevice": device_name,
                "TagName": _safe_get(tag, "Name"),
                "DataType": _safe_get(tag, "DataType"),
                "Connection": _safe_get(tag, "Connection"),
                "Address": _safe_get(tag, "Address"),
                "Comment": _safe_get(tag, "Comment"),
                "Length": _safe_get(tag, "Length"),
            })
    except Exception:
        pass
    return tags


def _collect_hmi_tags_from_table(tag_table, device_name, hmi_name):
    """通过 TagTable 收集变量"""
    tags = []
    try:
        for tag in tag_table.Tags:
            tags.append({
                "HmiDevice": device_name,
                "TagName": _safe_get(tag, "Name"),
                "DataType": _safe_get(tag, "DataType"),
                "Connection": _safe_get(tag, "Connection"),
                "Address": _safe_get(tag, "Address"),
                "Comment": _safe_get(tag, "Comment"),
                "Length": _safe_get(tag, "Length"),
            })
    except Exception:
        pass
    return tags


def export_device_io(project):
    """
    导出设备 I/O 映射和设备概览。

    返回: list[dict]  包含 Device, Slot, Module, IAddress, QAddress, Comment
    """
    io_list = []

    for device in project.Devices:
        try:
            device_name = device.Name
        except Exception:
            device_name = "(未知)"

        print(f"\n--- 设备 I/O: {device_name} ---")

        for device_item in device.DeviceItems:
            try:
                item_name = device_item.Name
            except Exception:
                item_name = "(未知)"

            # 获取硬件容器
            try:
                hw_container = device_item.GetService("HardwareContainer")
                if hw_container is not None:
                    for module in hw_container.Modules:
                        _extract_module_io(module, io_list, device_name, item_name)
            except Exception:
                pass

    return io_list


def _extract_module_io(module, io_list, device_name, device_item_name):
    """提取模块的 I/O 地址信息"""
    try:
        module_name = module.Name
        position = _safe_get(module, "PositionNumber", "")
    except Exception:
        return

    try:
        for io in module.IOs:
            io_list.append({
                "Device": device_name,
                "DeviceItem": device_item_name,
                "Module": module_name,
                "Slot": position,
                "IO_Name": _safe_get(io, "Name"),
                "IAddress": _safe_get(io, "IAddress"),
                "QAddress": _safe_get(io, "QAddress"),
                "Comment": _safe_get(io, "Comment"),
            })
    except Exception:
        pass


# ════════════════════════════════════════════════════════════
#  辅助函数
# ════════════════════════════════════════════════════════════

def _safe_get(obj, attr, default=""):
    """安全获取对象属性，异常时返回默认值"""
    try:
        val = getattr(obj, attr, default)
        if val is None:
            return default
        return str(val) if not isinstance(val, (int, float)) else val
    except Exception:
        return default


# ════════════════════════════════════════════════════════════
#  导出到 Excel
# ════════════════════════════════════════════════════════════

def write_to_excel(data_sheets, output_path):
    """
    将多个数据表写入一个 Excel 文件的不同 Sheet。

    data_sheets: dict  { sheet_name: (headers_list, data_list_of_dicts) }
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    for sheet_name, (headers, data) in data_sheets.items():
        if not data:
            continue

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet 名称最长 31 字符

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        # 自动调整列宽
        for col_idx, header in enumerate(headers, 1):
            max_width = max(
                len(str(row_data.get(header, ""))) if row_data.get(header) else 0
                for row_data in data
            )
            max_width = max(max_width, len(header)) + 2
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_width, 60)

    wb.save(str(output_path))
    print(f"\n[✓] 导出完成: {output_path}")


# ════════════════════════════════════════════════════════════
#  主函数
# ════════════════════════════════════════════════════════════

def main():
    """主流程"""
    print("=" * 60)
    print("  TIA Portal 变量批量导出工具 (Openness API)")
    print("  操作模式: 只读 — 不修改任何项目数据")
    print("=" * 60)
    print()

    # 1. 检查依赖
    api_info = check_dependencies()
    if api_info is None:
        sys.exit(1)

    # 2. 加载环境
    print("\n[步骤 1] 加载 Siemens.Engineering...")
    try:
        Siemens, TiaPortal_cls, TiaPortalMode = setup_environment(api_info)
    except Exception as e:
        print(f"\n[✗] 加载失败: {e}")
        traceback.print_exc()
        sys.exit(1)

    # 3. 连接 TIA Portal
    print("\n[步骤 2] 连接到 TIA Portal...")
    tia_portal = connect_to_tia(Siemens, TiaPortal_cls, TiaPortalMode)

    # 4. 获取打开的项目
    print("\n[步骤 3] 获取打开的项目...")
    project = get_open_project(tia_portal)
    try:
        project_path = project.Path
        project_name = Path(project_path).name if project_path else "(未知)"
    except Exception:
        project_name = "(未知)"
    print(f"[✓] 项目: {project_name}")

    # 5. 导出各类变量
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    data_sheets = {}

    # 5a. PLC 变量
    print("\n" + "=" * 60)
    print("[步骤 4a] 导出 PLC 变量表...")
    try:
        plc_tags = export_plc_tags(project)
        print(f"  共计 {len(plc_tags)} 个 PLC 变量")
        if plc_tags:
            headers = list(plc_tags[0].keys())
            data_sheets["PLC变量表"] = (headers, plc_tags)
    except Exception as e:
        print(f"  [!] PLC 变量导出失败: {e}")
        traceback.print_exc()

    # 5b. DB 块结构
    print("\n[步骤 4b] 导出 DB 块结构...")
    try:
        db_vars = export_db_blocks(project)
        print(f"  共计 {len(db_vars)} 个 DB 变量")
        if db_vars:
            headers = list(db_vars[0].keys())
            data_sheets["DB块结构"] = (headers, db_vars)
    except Exception as e:
        print(f"  [!] DB 块导出失败: {e}")
        traceback.print_exc()

    # 5c. HMI 变量
    print("\n[步骤 4c] 导出 HMI 变量...")
    try:
        hmi_tags = export_hmi_tags(project)
        print(f"  共计 {len(hmi_tags)} 个 HMI 变量")
        if hmi_tags:
            headers = list(hmi_tags[0].keys())
            data_sheets["HMI变量"] = (headers, hmi_tags)
    except Exception as e:
        print(f"  [!] HMI 变量导出失败: {e}")
        traceback.print_exc()

    # 5d. 设备 I/O
    print("\n[步骤 4d] 导出设备 I/O 映射...")
    try:
        io_mapping = export_device_io(project)
        print(f"  共计 {len(io_mapping)} 个 I/O 条目")
        if io_mapping:
            headers = list(io_mapping[0].keys())
            data_sheets["IO映射"] = (headers, io_mapping)
    except Exception as e:
        print(f"  [!] I/O 映射导出失败: {e}")
        traceback.print_exc()

    # 6. 写入 Excel
    if not data_sheets:
        print("\n[!] 没有导出任何数据。请检查 TIA Portal 中项目是否完整。")
        sys.exit(1)

    output_path = OUTPUT_DIR / f"TIA_Variables_{project_name}_{timestamp}.xlsx"
    print(f"\n[步骤 5] 写入 Excel: {output_path}")
    write_to_excel(data_sheets, output_path)

    # 7. 汇总
    print("\n" + "=" * 60)
    print("  导出汇总")
    print("=" * 60)
    total = 0
    for sheet_name, (headers, data) in data_sheets.items():
        count = len(data)
        total += count
        print(f"  {sheet_name}: {count} 条记录")
    print(f"  ─────────────────")
    print(f"  总计: {total} 条记录")
    print(f"  输出文件: {output_path}")
    print("\n[✓] 全部完成！")


if __name__ == "__main__":
    main()
