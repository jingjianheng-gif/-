#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""TIA Portal Openness API - 导出 e1_ 变量到 Excel"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import pythonnet
pythonnet.load('netfx')
import clr
import System
import os

API = r'C:\Program Files\Siemens\Automation\Portal V17\PublicAPI\V17'
System.Reflection.Assembly.LoadFrom(os.path.join(API, 'Siemens.Engineering.dll'))
clr.AddReference('Siemens.Engineering')

from Siemens.Engineering import TiaPortal, TiaPortalMode

print('连接 TIA Portal...')
tia = TiaPortal(TiaPortalMode.WithUserInterface)
projects = list(tia.Projects)
if not projects:
    print('ERROR: 未找到打开的项目')
    sys.exit(1)

proj = projects[0]
print(f'项目: {proj.Path}')

# 获取所有设备
all_tags = []
for device in proj.Devices:
    dname = device.Name
    for di in device.DeviceItems:
        try:
            sc = di.GetService('SoftwareContainer')
            if not sc or not sc.Software:
                continue
            sw = sc.Software
            ttg = sw.TagTableGroup
            # 递归遍历变量表
            def walk(group, prefix=''):
                try:
                    for sg in group.Groups:
                        walk(sg, prefix + '  ')
                except: pass
                try:
                    for tt in group.TagTables:
                        tname = tt.Name
                        for tag in tt.Tags:
                            tagname = str(tag.Name)
                            if tagname.startswith('e1_'):
                                addr = str(tag.LogicalAddress) if hasattr(tag,'LogicalAddress') else ''
                                dtype = str(tag.DataType) if hasattr(tag,'DataType') else ''
                                comment = str(tag.Comment) if hasattr(tag,'Comment') else ''
                                all_tags.append([tagname, dtype, addr, comment, tname, dname])
                except Exception as ex:
                    print(f'  ERROR reading tags: {ex}')
            walk(ttg)
        except:
            pass

print(f'\n找到 {len(all_tags)} 个 e1_ 变量')

if not all_tags:
    print('未找到 e1_ 变量。尝试列出所有变量名...')
    # 回退：列出前20个变量
    for device in proj.Devices:
        for di in device.DeviceItems:
            try:
                sc = di.GetService('SoftwareContainer')
                if sc and sc.Software:
                    for tt in sc.Software.TagTableGroup.TagTables:
                        for tag in tt.Tags:
                            name = str(tag.Name)
                            if not name.startswith('e1_'): continue
                            all_tags.append([name, str(tag.DataType), str(tag.LogicalAddress), '', str(tt.Name), str(device.Name)])
            except: pass
            if all_tags: break

print(f'最终: {len(all_tags)} 个变量')

# 生成 Excel
if all_tags:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'e1_extruder'
    hf = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF')
    for ci, h in enumerate(['Name','DataType','Address','Comment','TagTable','Device'], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hf
    for ri, row in enumerate(all_tags, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    for i in range(1,7):
        ws.column_dimensions[chr(64+i)].width = [28,14,18,30,18,18][i-1]
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{len(all_tags)+1}'

    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '导出结果')
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, '一号挤出机_e1_变量.xlsx')
    wb.save(out)
    print(f'\nDone! {out}')
else:
    print('\n导出失败。请确认 TIA Portal 项目已打开且包含 PLC 变量表。')
